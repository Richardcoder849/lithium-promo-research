#!/usr/bin/env python3
"""
生成锂电设备推广平台调研Excel看板报告
"""

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise ImportError(f"错误: 需要安装 openpyxl - {e}")


# 样式定义
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
S_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
A_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
B_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红
C_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 灰色

GRADE_FILLS = {
    "S": S_FILL,
    "A": A_FILL,
    "B": B_FILL,
    "C": C_FILL,
}


def load_platform_data(data_file):
    """加载平台数据"""
    with open(data_file, "r", encoding="utf-8") as f:
        return json.load(f)


def create_overview_sheet(wb, platforms):
    """Sheet 1: 平台总览"""
    ws = wb.active
    ws.title = "平台总览"
    
    headers = ["国家", "平台名称", "类型", "推荐等级", "影响力", "匹配产品线", "价格区间", "官网链接", "受众类型", "语言支持"]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入数据
    for row, p in enumerate(platforms, 2):
        ws.cell(row=row, column=1, value=p.get("country", ""))
        ws.cell(row=row, column=2, value=p.get("platform_name", ""))
        ws.cell(row=row, column=3, value=p.get("platform_type", ""))
        
        grade_cell = ws.cell(row=row, column=4, value=p.get("grade", ""))
        grade = p.get("grade", "")
        if grade in GRADE_FILLS:
            grade_cell.fill = GRADE_FILLS[grade]
        
        ws.cell(row=row, column=5, value=p.get("influence_score", ""))
        ws.cell(row=row, column=6, value=", ".join(p.get("matching_products", [])))
        ws.cell(row=row, column=7, value=p.get("price_range", ""))
        ws.cell(row=row, column=8, value=p.get("website", ""))
        ws.cell(row=row, column=9, value=p.get("audience_type", ""))
        ws.cell(row=row, column=10, value=p.get("language", ""))
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["H"].width = 35
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    return ws


def create_exhibition_sheet(wb, platforms):
    """Sheet 2: 展会详情"""
    ws = wb.create_sheet("展会详情")
    
    headers = ["展会名称", "国家", "举办时间", "展会规模", "参展费用", "联系方式", "备注", "官网链接"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    exhibitions = [p for p in platforms if p.get("platform_type") == "exhibition"]
    
    for row, p in enumerate(exhibitions, 2):
        ws.cell(row=row, column=1, value=p.get("platform_name", ""))
        ws.cell(row=row, column=2, value=p.get("country", ""))
        ws.cell(row=row, column=3, value=p.get("exhibition_time", ""))
        ws.cell(row=row, column=4, value=p.get("exhibition_scale", ""))
        ws.cell(row=row, column=5, value=p.get("price_range", ""))
        ws.cell(row=row, column=6, value=p.get("contact_info", ""))
        ws.cell(row=row, column=7, value=p.get("notes", ""))
        ws.cell(row=row, column=8, value=p.get("website", ""))
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["H"].width = 35
    
    ws.freeze_panes = "A2"
    return ws


def create_b2b_sheet(wb, platforms):
    """Sheet 3: B2B平台详情"""
    ws = wb.create_sheet("B2B平台")
    
    headers = ["平台名称", "国家", "入驻费用", "佣金比例", "主要受众", "语言支持", "链接", "备注"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    b2bs = [p for p in platforms if p.get("platform_type") == "b2b"]
    
    for row, p in enumerate(b2bs, 2):
        ws.cell(row=row, column=1, value=p.get("platform_name", ""))
        ws.cell(row=row, column=2, value=p.get("country", ""))
        ws.cell(row=row, column=3, value=p.get("entry_fee", ""))
        ws.cell(row=row, column=4, value=p.get("commission", ""))
        ws.cell(row=row, column=5, value=p.get("audience_type", ""))
        ws.cell(row=row, column=6, value=p.get("language", ""))
        ws.cell(row=row, column=7, value=p.get("website", ""))
        ws.cell(row=row, column=8, value=p.get("notes", ""))
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["G"].width = 35
    
    ws.freeze_panes = "A2"
    return ws


def create_media_sheet(wb, platforms):
    """Sheet 4: 媒体资源"""
    ws = wb.create_sheet("媒体资源")
    
    headers = ["媒体名称", "国家", "媒体类型", "广告价格", "发行量/流量", "投稿方式", "链接", "备注"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    medias = [p for p in platforms if p.get("platform_type") == "media"]
    
    for row, p in enumerate(medias, 2):
        ws.cell(row=row, column=1, value=p.get("platform_name", ""))
        ws.cell(row=row, column=2, value=p.get("country", ""))
        ws.cell(row=row, column=3, value=p.get("media_type", ""))
        ws.cell(row=row, column=4, value=p.get("ad_price", ""))
        ws.cell(row=row, column=5, value=p.get("circulation", ""))
        ws.cell(row=row, column=6, value=p.get("submission_method", ""))
        ws.cell(row=row, column=7, value=p.get("website", ""))
        ws.cell(row=row, column=8, value=p.get("notes", ""))
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["G"].width = 35
    
    ws.freeze_panes = "A2"
    return ws


def create_product_recommendation_sheet(wb, platforms, product_lines):
    """Sheet 5: 按产品线推荐"""
    ws = wb.create_sheet("产品线推荐")
    
    headers = ["产品线", "推荐平台", "平台类型", "国家", "推荐等级", "推荐理由"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    row = 2
    for product in product_lines:
        product_name = product.get("name", "")
        
        # 找到匹配该产品的平台
        matching = [p for p in platforms if product_name in p.get("matching_products", [])]
        matching.sort(key=lambda x: x.get("score", 0), reverse=True)
        
        for p in matching[:5]:  # 每个产品取Top5
            ws.cell(row=row, column=1, value=product_name)
            ws.cell(row=row, column=2, value=p.get("platform_name", ""))
            ws.cell(row=row, column=3, value=p.get("platform_type", ""))
            ws.cell(row=row, column=4, value=p.get("country", ""))
            
            grade_cell = ws.cell(row=row, column=5, value=p.get("grade", ""))
            grade = p.get("grade", "")
            if grade in GRADE_FILLS:
                grade_cell.fill = GRADE_FILLS[grade]
            
            ws.cell(row=row, column=6, value=p.get("recommend_reason", ""))
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["F"].width = 40
    
    ws.freeze_panes = "A2"
    return ws


def generate_excel_report(data_file, output_file=None, product_lines_file=None):
    """生成Excel看板报告"""
    
    # 加载平台数据
    platforms_data = load_platform_data(data_file)
    platforms = platforms_data.get("platforms", [])
    
    # 加载产品线
    product_lines = []
    if product_lines_file and Path(product_lines_file).exists():
        with open(product_lines_file, "r", encoding="utf-8") as f:
            product_lines = json.load(f).get("product_lines", [])
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 创建各Sheet
    create_overview_sheet(wb, platforms)
    create_exhibition_sheet(wb, platforms)
    create_b2b_sheet(wb, platforms)
    create_media_sheet(wb, platforms)
    create_product_recommendation_sheet(wb, platforms, product_lines)
    
    # 保存
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = Path(__file__).parent.parent / "data" / "reports"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = output_dir / f"promo_report_{timestamp}.xlsx"
    else:
        output_file = Path(output_file)
    
    wb.save(output_file)
    
    print(f"Excel看板报告已生成: {output_file}")
    print(f"包含 {len(platforms)} 个平台")
    
    # 统计
    stats = {
        "total": len(platforms),
        "by_country": {},
        "by_type": {},
        "by_grade": {},
    }
    
    for p in platforms:
        country = p.get("country", "Unknown")
        ptype = p.get("platform_type", "Unknown")
        grade = p.get("grade", "Unknown")
        
        stats["by_country"][country] = stats["by_country"].get(country, 0) + 1
        stats["by_type"][ptype] = stats["by_type"].get(ptype, 0) + 1
        stats["by_grade"][grade] = stats["by_grade"].get(grade, 0) + 1
    
    print("\n统计:")
    print(f"  总计: {stats['total']} 个平台")
    print(f"  按国家: {stats['by_country']}")
    print(f"  按类型: {stats['by_type']}")
    print(f"  按等级: {stats['by_grade']}")
    
    return output_file


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="生成推广平台Excel看板")
    parser.add_argument("data", help="平台数据JSON文件")
    parser.add_argument("--output", default=None, help="输出Excel文件路径")
    parser.add_argument("--products", default=None, help="产品线JSON文件路径")
    
    args = parser.parse_args()
    
    if not Path(args.data).exists():
        print(f"错误: 数据文件不存在: {args.data}")
        return 1
    
    # 默认产品线文件
    if args.products is None:
        default_products = Path(__file__).parent.parent / "references" / "product-lines.json"
        if default_products.exists():
            args.products = str(default_products)
    
    generate_excel_report(args.data, args.output, args.products)
    return 0


if __name__ == "__main__":
    sys.exit(main())
